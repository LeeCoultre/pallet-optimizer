"""SKU dimensions — source of truth for L×B×H + weight per Einheit.

The Einzelne-SKU distributor on the frontend looks these up in batch
during Pruefen to compute exact carton volumes/weights. Admins upload
a Dimensional_list.xlsx via the admin panel.

Lookup waterfall: fnsku → sku → ean. The first key that hits wins.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
import sqlalchemy as sa
from sqlalchemy import func, or_, select, text
from sqlalchemy.dialects.postgresql import ARRAY
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import get_db
from backend.deps import get_current_user, require_admin
from backend.orm import SkuDimension, User
from backend.schemas import (
    Paginated,
    SkuDimensionImportResult,
    SkuDimensionLookup,
    SkuDimensionLookupResponse,
    SkuDimensionRead,
    SkuDimensionUpsert,
)

router = APIRouter(prefix="/api", tags=["sku-dimensions"])


# ─── Lookup (any signed-in user) ─────────────────────────────────────

@router.get("/sku-dimensions/lookup", response_model=SkuDimensionLookupResponse)
async def lookup_sku_dimensions(
    keys: list[str] = Query(default_factory=list, description="FNSKU/SKU/EAN — repeat for batch"),
    _user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> SkuDimensionLookupResponse:
    """Batch lookup for FNSKU/SKU/EAN. Returns one entry per matched key.

    The same row can be reachable through multiple keys (multiple
    FNSKUs / SKUs / EANs in their respective arrays), so callers can
    probe with whichever identifier they have. Each requested key that
    hits gets bound under itself in the response."""
    keys_clean = [k.strip() for k in keys if k and k.strip()]
    if not keys_clean:
        return SkuDimensionLookupResponse()

    # Postgres array overlap: { fnskus && :keys } returns true if any
    # element of fnskus is in keys. Three OR clauses cover all key types.
    keys_param = sa.cast(keys_clean, ARRAY(sa.String))
    rows = (
        await db.execute(
            select(SkuDimension).where(
                or_(
                    SkuDimension.fnskus.op('&&')(keys_param),
                    SkuDimension.skus.op('&&')(keys_param),
                    SkuDimension.eans.op('&&')(keys_param),
                )
            )
        )
    ).scalars().all()

    lookups: dict[str, SkuDimensionLookup] = {}
    keys_set = set(keys_clean)
    for r in rows:
        compact = SkuDimensionLookup(
            length_cm=r.length_cm,
            width_cm=r.width_cm,
            height_cm=r.height_cm,
            weight_kg=r.weight_kg,
            source=r.source,
        )
        # Bind under EVERY one of the row's keys that the caller asked for
        for k in (*r.fnskus, *r.skus, *r.eans):
            if k in keys_set and k not in lookups:
                lookups[k] = compact

    missing = [k for k in keys_clean if k not in lookups]
    return SkuDimensionLookupResponse(lookups=lookups, missing=missing)


# ─── Admin: list / patch / delete ────────────────────────────────────

@router.get(
    "/admin/sku-dimensions",
    response_model=Paginated[SkuDimensionRead],
    dependencies=[Depends(require_admin)],
)
async def admin_list_sku_dimensions(
    q: Optional[str] = Query(None, description="Search FNSKU/SKU/EAN/Title"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    base = select(SkuDimension)
    if q:
        needle = q.strip()
        like = f"%{needle}%"
        # array_to_string lets us ILIKE across the whole array as one
        # haystack — substring match works for partial codes too.
        base = base.where(
            or_(
                func.array_to_string(SkuDimension.fnskus, ',').ilike(like),
                func.array_to_string(SkuDimension.skus, ',').ilike(like),
                func.array_to_string(SkuDimension.eans, ',').ilike(like),
                SkuDimension.title.ilike(like),
            )
        )

    total = (
        await db.execute(select(func.count()).select_from(base.subquery()))
    ).scalar_one()
    rows = (
        await db.execute(
            base.order_by(SkuDimension.updated_at.desc(), SkuDimension.id.desc())
            .limit(limit)
            .offset(offset)
        )
    ).scalars().all()
    return Paginated(
        items=[SkuDimensionRead.model_validate(r) for r in rows],
        total=total,
        limit=limit,
        offset=offset,
    )


def _dedupe(items: list[str]) -> list[str]:
    """Trim, drop blanks, preserve insertion order, dedupe."""
    seen, out = set(), []
    for it in items or []:
        s = (it or "").strip()
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return out


@router.post(
    "/admin/sku-dimensions",
    response_model=SkuDimensionRead,
    status_code=status.HTTP_201_CREATED,
)
async def admin_create_sku_dimension(
    payload: SkuDimensionUpsert,
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    fnskus, skus, eans = _dedupe(payload.fnskus), _dedupe(payload.skus), _dedupe(payload.eans)
    if not (fnskus or skus or eans):
        raise HTTPException(400, "At least one of fnskus/skus/eans must be non-empty")
    row = SkuDimension(
        fnskus=fnskus,
        skus=skus,
        eans=eans,
        title=payload.title,
        length_cm=payload.length_cm,
        width_cm=payload.width_cm,
        height_cm=payload.height_cm,
        weight_kg=payload.weight_kg,
        source="manual",
        updated_by=admin.email,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return SkuDimensionRead.model_validate(row)


@router.patch(
    "/admin/sku-dimensions/{row_id}",
    response_model=SkuDimensionRead,
)
async def admin_update_sku_dimension(
    row_id: int,
    payload: SkuDimensionUpsert,
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    row = await db.get(SkuDimension, row_id)
    if row is None:
        raise HTTPException(404, "Not found")
    fnskus, skus, eans = _dedupe(payload.fnskus), _dedupe(payload.skus), _dedupe(payload.eans)
    if not (fnskus or skus or eans):
        raise HTTPException(400, "At least one of fnskus/skus/eans must be non-empty")
    row.fnskus = fnskus
    row.skus = skus
    row.eans = eans
    row.title = payload.title
    row.length_cm = payload.length_cm
    row.width_cm = payload.width_cm
    row.height_cm = payload.height_cm
    row.weight_kg = payload.weight_kg
    row.source = "manual"
    row.updated_by = admin.email
    await db.commit()
    await db.refresh(row)
    return SkuDimensionRead.model_validate(row)


@router.delete(
    "/admin/sku-dimensions/{row_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_admin)],
)
async def admin_delete_sku_dimension(
    row_id: int,
    db: AsyncSession = Depends(get_db),
):
    row = await db.get(SkuDimension, row_id)
    if row is None:
        raise HTTPException(404, "Not found")
    await db.delete(row)
    await db.commit()
    return None


# ─── Admin: xlsx import ──────────────────────────────────────────────
# Column aliases (lowercase). First column matching any alias wins.
_FNSKU_ALIASES = {"fnsku", "x-code", "asin-fnsku"}
_SKU_ALIASES = {"sku", "merchant sku", "seller sku"}
_EAN_ALIASES = {"ean", "upc", "gtin", "barcode", "ean/upc", "ean-code", "ean code"}
_TITLE_ALIASES = {"title", "name", "bezeichnung", "produkt", "artikel", "description"}
_L_ALIASES = {"l", "länge", "laenge", "length", "len", "l (cm)", "länge (cm)", "laenge (cm)"}
_B_ALIASES = {"b", "w", "breite", "width", "b (cm)", "breite (cm)"}
_H_ALIASES = {"h", "höhe", "hoehe", "height", "h (cm)", "höhe (cm)", "hoehe (cm)"}
_WEIGHT_ALIASES = {
    "gewicht", "weight", "kg", "gewicht (kg)", "weight (kg)",
    "gewicht_kg", "weight_kg", "masse",
}


def _find_col(headers: list[str], aliases: set[str]) -> Optional[int]:
    # Pass 1: exact match (lowercase)
    for i, h in enumerate(headers):
        norm = (h or "").strip().lower()
        if norm in aliases:
            return i
    # Pass 2: substring match — but ALIAS must be the substring of header,
    # not the other way around (avoids "L" matching "Pallet load")
    for i, h in enumerate(headers):
        norm = (h or "").strip().lower()
        for a in aliases:
            if a and a in norm:
                return i
    return None


def _parse_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _normalize_ean_str(s: str) -> Optional[str]:
    """Accept pure digits 8-14 chars; reject descriptive text."""
    s = s.strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    if s.isdigit() and 8 <= len(s) <= 14:
        return s
    return None


def _parse_keys(v, *, ean_only: bool = False) -> list[str]:
    """Cell may carry one key, several keys (comma/semicolon/newline-
    separated), or a number that Excel turned into a float. Returns a
    deduped list of normalised string keys."""
    if v is None or v == "":
        return []
    # Numeric arrival from Excel — single value
    if isinstance(v, float) and v.is_integer():
        s = str(int(v))
    elif isinstance(v, int):
        s = str(v)
    else:
        s = str(v)
    # Split on common separators and trim
    parts = [p.strip() for p in s.replace(";", ",").replace("\n", ",").split(",")]
    out, seen = [], set()
    for p in parts:
        if not p:
            continue
        if ean_only:
            norm = _normalize_ean_str(p)
            if norm is None:
                continue
            p = norm
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _str_or_none(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


@router.post(
    "/admin/sku-dimensions/import",
    response_model=SkuDimensionImportResult,
    dependencies=[Depends(require_admin)],
)
async def admin_import_sku_dimensions(
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
) -> SkuDimensionImportResult:
    """Upload Dimensional_list.xlsx. The same physical product often
    ships under several FNSKUs/SKUs (PRIME vs EV vs regional channel),
    so this importer DEDUPES: any incoming row that shares ANY key
    (FNSKU / SKU / EAN) with an existing row is merged into it — the
    keys accumulate, dimensions get refreshed.

    Required columns (case-insensitive, alias-tolerant):
      FNSKU | SKU | EAN | Title | L (cm) | B (cm) | H (cm) | Gewicht (kg)
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx / .xls files accepted")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot open file: {e}")

    ws = wb.active
    rows = list(ws.values)
    wb.close()

    # Locate the header row — first row with L+B+H+weight columns present
    header_idx = None
    cols: dict[str, Optional[int]] = {}
    for idx, row in enumerate(rows):
        if not row:
            continue
        headers = [str(c or "").strip().lower() for c in row]
        c_l = _find_col(headers, _L_ALIASES)
        c_b = _find_col(headers, _B_ALIASES)
        c_h = _find_col(headers, _H_ALIASES)
        c_w = _find_col(headers, _WEIGHT_ALIASES)
        if all(c is not None for c in (c_l, c_b, c_h, c_w)):
            header_idx = idx
            cols = {
                "fnsku": _find_col(headers, _FNSKU_ALIASES),
                "sku": _find_col(headers, _SKU_ALIASES),
                "ean": _find_col(headers, _EAN_ALIASES),
                "title": _find_col(headers, _TITLE_ALIASES),
                "l": c_l,
                "b": c_b,
                "h": c_h,
                "weight": c_w,
            }
            break

    if header_idx is None:
        raise HTTPException(
            422,
            "Could not detect header row. Required columns: L (cm), B (cm), "
            "H (cm), Gewicht (kg) — plus at least one of FNSKU/SKU/EAN.",
        )

    result = SkuDimensionImportResult()
    now = datetime.now(timezone.utc)

    # Pre-fetch all existing rows + build per-key indexes once. Same row
    # can sit in multiple maps (one per key element across all 3 arrays).
    existing_rows = (await db.execute(select(SkuDimension))).scalars().all()
    by_fnsku: dict[str, SkuDimension] = {}
    by_sku: dict[str, SkuDimension] = {}
    by_ean: dict[str, SkuDimension] = {}
    for r in existing_rows:
        for k in r.fnskus or []: by_fnsku[k] = r
        for k in r.skus or []:   by_sku[k] = r
        for k in r.eans or []:   by_ean[k] = r

    def add_key(target_map: dict, target_list: list, key: Optional[str], row: SkuDimension):
        if not key or key in target_list:
            return
        target_list.append(key)
        target_map[key] = row

    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row or all(c is None for c in row):
            continue

        def cell(key: str):
            i = cols.get(key)
            if i is None or i >= len(row):
                return None
            return row[i]

        fnskus_in = _parse_keys(cell("fnsku"))
        skus_in = _parse_keys(cell("sku"))
        eans_in = _parse_keys(cell("ean"), ean_only=True)
        title = _str_or_none(cell("title"))
        l_val = _parse_float(cell("l"))
        b_val = _parse_float(cell("b"))
        h_val = _parse_float(cell("h"))
        w_val = _parse_float(cell("weight"))

        if not (fnskus_in or skus_in or eans_in):
            result.skipped += 1
            result.warnings.append(f"Row {row_idx + 1}: no FNSKU/SKU/EAN — skipped")
            continue
        if any(v is None or v <= 0 for v in (l_val, b_val, h_val, w_val)):
            label = (fnskus_in or skus_in or eans_in)[0]
            result.skipped += 1
            result.warnings.append(
                f"Row {row_idx + 1} '{label}': missing or non-positive L/B/H/Gewicht — skipped"
            )
            continue

        # Find an existing row that shares ANY key (across all 3 arrays
        # in this incoming row). First hit wins; we don't try to merge
        # multiple matching rows together (admin should clean those up).
        existing = None
        for k in fnskus_in:
            if k in by_fnsku: existing = by_fnsku[k]; break
        if existing is None:
            for k in skus_in:
                if k in by_sku: existing = by_sku[k]; break
        if existing is None:
            for k in eans_in:
                if k in by_ean: existing = by_ean[k]; break

        if existing is not None:
            # Accumulate new keys into the appropriate arrays.
            # SQLAlchemy only flags ARRAY columns as changed if the list
            # identity changes, so we replace the list rather than mutate.
            for k in fnskus_in:
                if k not in (existing.fnskus or []):
                    existing.fnskus = [*(existing.fnskus or []), k]
                    by_fnsku[k] = existing
            for k in skus_in:
                if k not in (existing.skus or []):
                    existing.skus = [*(existing.skus or []), k]
                    by_sku[k] = existing
            for k in eans_in:
                if k not in (existing.eans or []):
                    existing.eans = [*(existing.eans or []), k]
                    by_ean[k] = existing
            existing.title = title or existing.title
            existing.length_cm = l_val
            existing.width_cm = b_val
            existing.height_cm = h_val
            existing.weight_kg = w_val
            existing.source = "xlsx_import"
            existing.updated_by = admin.email
            existing.updated_at = now
            result.updated += 1
        else:
            new_row = SkuDimension(
                fnskus=fnskus_in,
                skus=skus_in,
                eans=eans_in,
                title=title,
                length_cm=l_val,
                width_cm=b_val,
                height_cm=h_val,
                weight_kg=w_val,
                source="xlsx_import",
                updated_by=admin.email,
            )
            db.add(new_row)
            # Same-batch dedupe: a later row in the SAME xlsx that shares
            # any of these keys merges into this new row.
            for k in fnskus_in: by_fnsku[k] = new_row
            for k in skus_in:   by_sku[k] = new_row
            for k in eans_in:   by_ean[k] = new_row
            result.imported += 1

    await db.commit()
    return result


# ─── Admin: xlsx export ──────────────────────────────────────────────
# Mirror format of the upload, so a round-trip (export → re-import) is
# lossless. Lets the admin treat the DB as the source of truth without
# fear of losing data — they can always download a fresh xlsx backup.

@router.get(
    "/admin/sku-dimensions/export",
    dependencies=[Depends(require_admin)],
)
async def admin_export_sku_dimensions(
    db: AsyncSession = Depends(get_db),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    rows = (
        await db.execute(
            select(SkuDimension).order_by(SkuDimension.id)
        )
    ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dimensions"
    headers = ["FNSKU", "SKU", "EAN", "Title", "L (cm)", "B (cm)", "H (cm)", "Gewicht (kg)", "Source", "Updated"]
    ws.append(headers)
    # Bold header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Multi-key rows: comma-separate the lists. Re-import treats each
    # token as its own row, then merges back via dedupe — round-trip
    # lossless even though the cells display flat.
    for r in rows:
        ws.append([
            ", ".join(r.fnskus or []) or None,
            ", ".join(r.skus or []) or None,
            ", ".join(r.eans or []) or None,
            r.title,
            r.length_cm,
            r.width_cm,
            r.height_cm,
            r.weight_kg,
            r.source,
            r.updated_at.strftime("%Y-%m-%d %H:%M") if r.updated_at else None,
        ])
    # Auto column widths
    for col in ws.columns:
        m = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(40, m + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"dimensions_export_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
