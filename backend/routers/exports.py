"""xlsx exports — Berichte tab.

Open to all authenticated users (per product decision: warehouse team
shares one report). Generates an in-memory openpyxl workbook and
streams it back. Date filters apply to `finished_at` because exports
are about completed work.

If the result set is huge (>50k rows) we'd want a streaming xlsx
writer; for the current 5-user warehouse the in-memory build is fine
and orders of magnitude simpler.
"""

import io
from datetime import date, datetime, time, timezone
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database import get_db
from backend.deps import get_current_user
from backend.orm import Auftrag, AuftragStatus, User

router = APIRouter(prefix="/api/exports", tags=["exports"])


_HEADER_FILL = PatternFill("solid", fgColor="111111")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

_COLUMNS = [
    ("Datum",         18),
    ("Operator",      22),
    ("Sendungsnr.",   24),
    ("Datei",         34),
    ("Pal.",           7),
    ("Artikel",        9),
    ("Dauer (min)",   13),
    ("Status",        12),
]


def _fba_from_parsed(parsed: Optional[dict]) -> Optional[str]:
    if not parsed:
        return None
    meta = parsed.get("meta") or {}
    return meta.get("sendungsnummer") or meta.get("fbaCode")


@router.get("/auftraege.xlsx")
async def export_auftraege(
    from_: Optional[date] = Query(None, alias="from"),
    to: Optional[date] = Query(None),
    _me: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = (
        select(Auftrag, User)
        .join(User, Auftrag.assigned_to_user_id == User.id, isouter=True)
        .where(Auftrag.status == AuftragStatus.completed)
        .order_by(Auftrag.finished_at.desc())
    )

    if from_ is not None:
        q = q.where(Auftrag.finished_at >= datetime.combine(from_, time.min, tzinfo=timezone.utc))
    if to is not None:
        q = q.where(Auftrag.finished_at < datetime.combine(to, time.max, tzinfo=timezone.utc))

    rows = (await db.execute(q)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Aufträge"

    for col_idx, (label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"

    for r_idx, (auftrag, user) in enumerate(rows, start=2):
        parsed = auftrag.parsed or {}
        pallets = parsed.get("pallets") or []
        article_count = sum(len(p.get("items") or []) for p in pallets)
        finished = auftrag.finished_at
        finished_local = finished.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M") if finished else ""
        dauer_min = round((auftrag.duration_sec or 0) / 60, 1) if auftrag.duration_sec else None

        ws.cell(row=r_idx, column=1, value=finished_local)
        ws.cell(row=r_idx, column=2, value=user.name if user else "")
        ws.cell(row=r_idx, column=3, value=_fba_from_parsed(parsed) or "")
        ws.cell(row=r_idx, column=4, value=auftrag.file_name)
        ws.cell(row=r_idx, column=5, value=len(pallets))
        ws.cell(row=r_idx, column=6, value=article_count)
        ws.cell(row=r_idx, column=7, value=dauer_min)
        ws.cell(row=r_idx, column=8, value=auftrag.status.value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = date.today().isoformat()
    # Content-Disposition is latin-1 only — use ASCII ellipsis ('...').
    range_part = ""
    if from_ or to:
        range_part = f"_{from_ or '...'}_{to or '...'}"
    filename = f"marathon-auftraege_{today}{range_part}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            # Tell the frontend the row count so it can show "32 rows exported".
            "X-Row-Count": str(len(rows)),
        },
    )
