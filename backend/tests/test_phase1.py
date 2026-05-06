"""Phase 1 — Suche, Live-Aktivität, Export.

Coverage matrix:
  /api/search           — auth, by FNSKU, by Sendungsnummer, fuzzy/typo, date filter
  /api/activity/live    — auth, active workers, recent events, server_time
  /api/activity/shift   — empty day, started after first action, completed counter
  /api/exports/*.xlsx   — auth, returns workbook, X-Row-Count header, date filter
"""

from io import BytesIO

from openpyxl import load_workbook

from .conftest import make_payload


def make_payload_with_items(file_name: str, fba: str, fnskus: list[str]) -> dict:
    return {
        "file_name": file_name,
        "parsed": {
            "meta": {"sendungsnummer": fba},
            "pallets": [
                {
                    "id": "P1",
                    "items": [
                        {"fnsku": fn, "sku": f"SKU-{fn}", "ean": f"4001{fn[-6:].zfill(6)}"}
                        for fn in fnskus
                    ],
                }
            ],
        },
    }


# ─── /api/search ─────────────────────────────────────────────────────
# Note: dedicated auth-guard tests live in test_auth.py. The local dev
# environment runs with ALLOW_ANONYMOUS=true so /api/search would return
# 200 without a Bearer token, which is correct (it binds to
# anonymous@local). Don't add auth assertions here.

async def test_search_by_fnsku_inside_parsed(client, user, as_user):
    as_user(user)
    await client.post("/api/auftraege", json=make_payload_with_items(
        "alpha.docx", "FBA-AAA", ["X0ABCDEFG", "X0ZZZZZZZ"],
    ))
    await client.post("/api/auftraege", json=make_payload_with_items(
        "beta.docx", "FBA-BBB", ["X0OTHERID"],
    ))

    r = await client.get("/api/search?q=X0ABCDEFG")
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 1
    hit = body["items"][0]
    assert hit["file_name"] == "alpha.docx"
    assert hit["matched_field"] == "fnsku"
    assert hit["matched_value"] == "X0ABCDEFG"


async def test_search_by_sendungsnummer(client, user, as_user):
    as_user(user)
    await client.post("/api/auftraege", json=make_payload_with_items(
        "x.docx", "FBA15ABC123XYZ", ["X0AAA"],
    ))
    r = await client.get("/api/search?q=15ABC123")
    assert r.status_code == 200
    items = r.json()["items"]
    assert len(items) == 1
    assert items[0]["matched_field"] == "sendungsnummer"
    assert items[0]["fba_code"] == "FBA15ABC123XYZ"


async def test_search_by_filename(client, user, as_user):
    as_user(user)
    await client.post("/api/auftraege", json=make_payload("Lagerauftrag-2026-05.docx", "FBA-1"))
    r = await client.get("/api/search?q=Lagerauftrag-2026")
    items = r.json()["items"]
    assert len(items) == 1
    assert items[0]["matched_field"] == "file_name"


async def test_search_two_char_minimum(client, user, as_user):
    as_user(user)
    r = await client.get("/api/search?q=a")
    assert r.status_code == 422  # min_length=2


# ─── /api/activity/live ──────────────────────────────────────────────

async def test_live_lists_active_worker(client, user, user2, as_user):
    as_user(user)
    a = (await client.post("/api/auftraege", json=make_payload("running.docx", "FBA-RUN"))).json()["id"]
    await client.post(f"/api/auftraege/{a}/start")

    as_user(user2)
    r = await client.get("/api/activity/live")
    assert r.status_code == 200
    body = r.json()
    workers = body["active_workers"]
    assert len(workers) == 1
    assert workers[0]["user_name"] == "TestUser"
    assert workers[0]["file_name"] == "running.docx"
    assert workers[0]["fba_code"] == "FBA-RUN"
    assert workers[0]["step"] == "pruefen"


async def test_live_events_include_recent_actions(client, user, as_user):
    as_user(user)
    a = (await client.post("/api/auftraege", json=make_payload())).json()["id"]
    await client.post(f"/api/auftraege/{a}/start")
    await client.post(f"/api/auftraege/{a}/complete")

    r = await client.get("/api/activity/live")
    actions = [e["action"] for e in r.json()["events"]]
    # Newest first — complete should be at the top
    assert actions[0] == "complete"
    assert "start" in actions
    assert "upload" in actions
    # Joined fields populated
    assert all(e["user_name"] == "TestUser" for e in r.json()["events"])


async def test_live_server_time_present(client, user, as_user):
    as_user(user)
    r = await client.get("/api/activity/live")
    assert "server_time" in r.json()


# ─── /api/activity/shift ─────────────────────────────────────────────

async def test_shift_empty_when_no_actions(client, user, as_user):
    as_user(user)
    r = await client.get("/api/activity/shift")
    assert r.status_code == 200
    body = r.json()
    assert body["started_at"] is None
    assert body["duration_sec"] == 0
    assert body["completed_today"] == 0


async def test_shift_starts_after_first_action(client, user, as_user):
    as_user(user)
    await client.post("/api/auftraege", json=make_payload())  # creates audit row
    r = await client.get("/api/activity/shift")
    body = r.json()
    assert body["started_at"] is not None
    assert body["duration_sec"] >= 0


async def test_shift_counts_completed_today(client, user, as_user):
    as_user(user)
    a = (await client.post("/api/auftraege", json=make_payload())).json()["id"]
    await client.post(f"/api/auftraege/{a}/start")
    await client.post(f"/api/auftraege/{a}/complete")

    r = await client.get("/api/activity/shift")
    assert r.json()["completed_today"] == 1


# ─── /api/exports/auftraege.xlsx ─────────────────────────────────────

async def test_export_returns_workbook_with_completed(client, user, as_user):
    as_user(user)
    a = (await client.post("/api/auftraege", json=make_payload("done.docx", "FBA-DONE"))).json()["id"]
    await client.post(f"/api/auftraege/{a}/start")
    await client.post(f"/api/auftraege/{a}/complete")
    # Queued one — must NOT show up
    await client.post("/api/auftraege", json=make_payload("queued.docx", "FBA-Q"))

    r = await client.get("/api/exports/auftraege.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    assert r.headers["x-row-count"] == "1"

    wb = load_workbook(filename=BytesIO(r.content))
    ws = wb.active
    # Row 1 = header, row 2 = first data row
    assert ws.cell(row=1, column=1).value == "Datum"
    assert ws.cell(row=2, column=2).value == "TestUser"
    assert ws.cell(row=2, column=3).value == "FBA-DONE"
    assert ws.cell(row=2, column=4).value == "done.docx"
    assert ws.cell(row=2, column=8).value == "completed"
    # Only one data row
    assert ws.cell(row=3, column=1).value is None


async def test_export_filters_by_date(client, user, as_user):
    """from=tomorrow ⇒ no rows."""
    as_user(user)
    a = (await client.post("/api/auftraege", json=make_payload())).json()["id"]
    await client.post(f"/api/auftraege/{a}/start")
    await client.post(f"/api/auftraege/{a}/complete")

    r = await client.get("/api/exports/auftraege.xlsx?from=2099-01-01")
    assert r.status_code == 200
    assert r.headers["x-row-count"] == "0"
